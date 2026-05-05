import os, io, json, sqlite3, re
from datetime import datetime, date, timedelta
from functools import wraps
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import (Flask, render_template, request, jsonify,
                   send_file, redirect, url_for, session, flash, g)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'daksham-capital-2026-secret')
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('RAILWAY_ENVIRONMENT') is not None
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours
DB_PATH   = os.path.join(os.path.dirname(__file__), 'data', 'daksham.db')
ADMIN_PWD = os.environ.get('ADMIN_PASSWORD', 'Admin@123')
DASH_USER = os.environ.get('DASH_USERNAME', 'DakshamEmployee')
DASH_PWD  = os.environ.get('DASH_PASSWORD', 'Daksham@2026')

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    db.executescript("""
    CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT UNIQUE NOT NULL, name TEXT NOT NULL,
        universe TEXT NOT NULL DEFAULT 'mf', parent_group TEXT NOT NULL DEFAULT 'equity',
        display_order INTEGER DEFAULT 0, is_active INTEGER DEFAULT 1, bm_type TEXT DEFAULT 'explicit'
    );
    CREATE TABLE IF NOT EXISTS funds (
        id INTEGER PRIMARY KEY AUTOINCREMENT, category_id INTEGER NOT NULL, name TEXT NOT NULL,
        short_name TEXT, type TEXT DEFAULT 'fund', inception_date TEXT,
        is_active INTEGER DEFAULT 1, qualifies_rolling INTEGER DEFAULT 0,
        FOREIGN KEY (category_id) REFERENCES categories(id)
    );
    CREATE TABLE IF NOT EXISTS fund_details (
        fund_id INTEGER PRIMARY KEY, manager1 TEXT, manager2 TEXT,
        aum_latest REAL, expense_ratio REAL, std_dev REAL, beta REAL, info_ratio REAL,
        sharpe REAL, volatility REAL,
        lc_pct REAL, mc_pct REAL, sc_pct REAL, cash_pct REAL,
        sector1_name TEXT, sector1_pct REAL, sector2_name TEXT, sector2_pct REAL,
        sector3_name TEXT, sector3_pct REAL,
        equity_pct REAL, net_equity_pct REAL, debt_pct REAL, others_pct REAL,
        gold_pct REAL, silver_pct REAL, reits_pct REAL,
        updated_at TEXT,
        FOREIGN KEY (fund_id) REFERENCES funds(id)
    );
    CREATE TABLE IF NOT EXISTS fund_details_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fund_id INTEGER NOT NULL, quarter TEXT NOT NULL,
        aum REAL, lc_pct REAL, mc_pct REAL, sc_pct REAL,
        sector1_name TEXT, sector1_pct REAL, sector2_name TEXT, sector2_pct REAL,
        sector3_name TEXT, sector3_pct REAL,
        equity_pct REAL, net_equity_pct REAL, debt_pct REAL, others_pct REAL,
        gold_pct REAL, silver_pct REAL, reits_pct REAL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (fund_id) REFERENCES funds(id), UNIQUE(fund_id, quarter)
    );
    CREATE TABLE IF NOT EXISTS navs (
        fund_id INTEGER NOT NULL, nav_date TEXT NOT NULL, nav_value REAL NOT NULL,
        PRIMARY KEY (fund_id, nav_date), FOREIGN KEY (fund_id) REFERENCES funds(id)
    );
    CREATE TABLE IF NOT EXISTS precalc_returns (
        fund_id INTEGER NOT NULL, as_of_date TEXT NOT NULL,
        ytd REAL, ret_1m REAL, ret_3m REAL, ret_6m REAL,
        ret_1y REAL, ret_2y REAL, ret_3y REAL, ret_5y REAL, ret_7y REAL, ret_10y REAL,
        since_inception REAL, since_inception_bm REAL, roll_3y REAL, roll_5y REAL,
        outperf_3y REAL, outperf_5y REAL,
        cy2027 REAL, cy2026 REAL, cy2025 REAL, cy2024 REAL, cy2023 REAL, cy2022 REAL,
        cy2021 REAL, cy2020 REAL, cy2019 REAL, cy2018 REAL, cy2017 REAL, cy2016 REAL,
        cy2015 REAL, cy2014 REAL, cy2013 REAL, cy2012 REAL, cy2011 REAL, cy2010 REAL,
        cy2009 REAL, cy2008 REAL, cy2007 REAL, cy2006 REAL,
        PRIMARY KEY (fund_id, as_of_date)
    );
    CREATE INDEX IF NOT EXISTS idx_navs_fund_date ON navs(fund_id, nav_date);
    CREATE INDEX IF NOT EXISTS idx_navs_date ON navs(nav_date);
    CREATE INDEX IF NOT EXISTS idx_funds_category ON funds(category_id);
    CREATE INDEX IF NOT EXISTS idx_precalc ON precalc_returns(fund_id, as_of_date);
    """)
    existing = db.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
    if existing == 0:
        cats = [
            # MF - Equity
            ('mf_largecap',  'Large Cap',                     'mf','equity',  1,'explicit'),
            ('mf_midcap',    'Mid Cap',                       'mf','equity',  2,'explicit'),
            ('mf_smallcap',  'Small Cap',                     'mf','equity',  3,'explicit'),
            ('mf_multicap',  'DC Multi Cap',                  'mf','equity',  4,'explicit'),
            ('mf_index',     'Index Funds',                   'mf','equity',  5,'none'),
            ('mf_others',    'Others - Mfg & Business Cycle', 'mf','equity',  6,'explicit'),
            # MF - Hybrids
            ('mf_baf',       'BAF',                           'mf','hybrids', 7,'explicit'),
            ('mf_maf',       'MAF',                           'mf','hybrids', 8,'explicit'),
            ('mf_agg_hybrid','Aggressive Hybrid',             'mf','hybrids', 9,'explicit'),
            ('mf_eq_savings','Equity Savings',                'mf','hybrids',10,'explicit'),
            # Non MF
            ('nmf_largecap', 'Large Caps',                    'nmf','non_mf',11,'explicit'),
            ('nmf_multicap', 'MultiCap',                      'nmf','non_mf',12,'explicit'),
            ('nmf_smids',    'SMIDs',                         'nmf','non_mf',13,'explicit'),
            # Indices
            ('idx_indices',  'Indices',                       'idx','indices',14,'none'),
        ]
        db.executemany("INSERT INTO categories(code,name,universe,parent_group,display_order,bm_type) VALUES(?,?,?,?,?,?)", cats)
    db.commit(); db.close()

init_db()

# ── AUTH ────────────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'): return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

def check_admin_pwd(pwd): return pwd == ADMIN_PWD

def dash_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('dash_logged_in'): return redirect(url_for('dash_login'))
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET','POST'])
def dash_login():
    if session.get('dash_logged_in'): return redirect(url_for('index'))
    if request.method == 'POST':
        if request.form.get('username','') == DASH_USER and request.form.get('password','') == DASH_PWD:
            session['dash_logged_in'] = True; return redirect(url_for('index'))
        flash('Incorrect username or password')
    return render_template('dash_login.html')

@app.route('/logout')
def dash_logout(): session.pop('dash_logged_in', None); return redirect(url_for('dash_login'))

# ── RETURN CALCULATIONS ────────────────────────────────────────────────────────
def load_series(db, fund_id):
    rows = db.execute("SELECT nav_date, nav_value FROM navs WHERE fund_id=? ORDER BY nav_date", (fund_id,)).fetchall()
    if not rows: return pd.Series(dtype=float)
    return pd.Series([r['nav_value'] for r in rows], index=pd.to_datetime([r['nav_date'] for r in rows]))

def get_month_end_navs(s):
    if len(s) == 0: return s
    return s.resample('ME').last().dropna()

def get_nav_at(s, target):
    if len(s) == 0: return None
    m = get_month_end_navs(s)
    sub = m[m.index <= pd.Timestamp(target)]
    return float(sub.iloc[-1]) if len(sub) > 0 else None

def calc_trailing(s, as_of, years=None, months=None):
    m = get_month_end_navs(s)
    if len(m) == 0: return None
    sub = m[m.index <= pd.Timestamp(as_of)]
    if len(sub) == 0: return None
    now_idx = len(sub) - 1
    now_val = float(sub.iloc[-1])
    n = (years * 12) if years else months
    target_idx = now_idx - n
    if target_idx < 0: return None
    then_val = float(sub.iloc[target_idx])
    if then_val == 0: return None
    if years:
        actual_days = (sub.index[now_idx] - sub.index[target_idx]).days
        if actual_days <= 0: return None
        return round(((now_val / then_val) ** (365 / actual_days) - 1) * 100, 4)
    else:
        return round((now_val / then_val - 1) * 100, 4)

def calc_ytd(s, as_of):
    m = get_month_end_navs(s)
    if len(m) == 0: return None
    sub = m[m.index <= pd.Timestamp(as_of)]
    if len(sub) == 0: return None
    now_val = float(sub.iloc[-1])
    yr = sub.index[-1].year
    dec_prev = m[m.index <= pd.Timestamp(f'{yr-1}-12-31')]
    if len(dec_prev) == 0: return None
    then_val = float(dec_prev.iloc[-1])
    if then_val == 0: return None
    return round((now_val / then_val - 1) * 100, 4)

def calc_cy(s, year, as_of):
    m = get_month_end_navs(s)
    if len(m) == 0: return None
    dec_prev = m[m.index <= pd.Timestamp(f'{year-1}-12-31')]
    if len(dec_prev) == 0: return None
    start_val = float(dec_prev.iloc[-1])
    if start_val == 0: return None
    end_target = as_of if year == pd.Timestamp(as_of).year else f'{year}-12-31'
    end_sub = m[m.index <= pd.Timestamp(end_target)]
    if len(end_sub) == 0: return None
    return round((float(end_sub.iloc[-1]) / start_val - 1) * 100, 4)

def calc_since_inc(s, as_of):
    first = s.dropna()
    if len(first) == 0: return None, None
    inc_date, inc_nav = first.index[0], float(first.iloc[0])
    now = get_nav_at(s, as_of)
    if not now or inc_nav == 0: return None, str(inc_date.date())
    yrs = (pd.Timestamp(as_of) - inc_date).days / 365.25
    if yrs < 0.5: return None, str(inc_date.date())
    return round(((now / inc_nav) ** (1 / yrs) - 1) * 100, 4), str(inc_date.date())

def calc_since_inc_bm(fund_s, bm_s, as_of):
    first = fund_s.dropna()
    if len(first) == 0 or len(bm_s) == 0: return None
    inc_date = first.index[0]
    bm_start = get_nav_at(bm_s, inc_date)
    bm_now = get_nav_at(bm_s, as_of)
    if not bm_start or not bm_now or bm_start == 0: return None
    yrs = (pd.Timestamp(as_of) - inc_date).days / 365.25
    if yrs < 0.5: return None
    return round(((bm_now / bm_start) ** (1 / yrs) - 1) * 100, 4)

def calc_custom(s, from_date, to_date):
    n1, n2 = get_nav_at(s, from_date), get_nav_at(s, to_date)
    if not n1 or not n2 or n1 == 0: return None, None, None
    days = (pd.Timestamp(to_date) - pd.Timestamp(from_date)).days
    if days <= 0: return None, None, None
    abs_ret = round((n2 / n1 - 1) * 100, 4)
    ann_ret = round(((n2 / n1) ** (365 / days) - 1) * 100, 4) if days >= 365 else None
    return abs_ret, ann_ret, days

def calc_rolling(s, years, as_of):
    m = get_month_end_navs(s)
    if len(m) == 0: return None
    as_of_ts = pd.Timestamp(as_of)
    ten_yr_start = as_of_ts - pd.DateOffset(years=10)
    m = m[(m.index >= ten_yr_start) & (m.index <= as_of_ts)]
    if len(m) == 0: return None
    n = years * 12
    vals = []
    for i in range(n, len(m)):
        en, sn = float(m.iloc[i]), float(m.iloc[i - n])
        if sn > 0:
            actual_days = (m.index[i] - m.index[i - n]).days
            if actual_days > 0:
                vals.append(((en / sn) ** (365 / actual_days) - 1) * 100)
    return round(np.mean(vals), 4) if vals else None

def calc_outperf(fund_s, bm_s, years, as_of, threshold=0.005):
    fm, bm_m = get_month_end_navs(fund_s), get_month_end_navs(bm_s)
    as_of_ts = pd.Timestamp(as_of)
    ten_yr_start = as_of_ts - pd.DateOffset(years=10)
    fm = fm[(fm.index >= ten_yr_start) & (fm.index <= as_of_ts)]
    bm_m = bm_m[(bm_m.index >= ten_yr_start) & (bm_m.index <= as_of_ts)]
    if len(fm) == 0 or len(bm_m) == 0: return None
    common = fm.index.intersection(bm_m.index)
    if len(common) == 0: return None
    n = years * 12
    fm_c = fm[fm.index.isin(common)]
    bm_c = bm_m[bm_m.index.isin(common)]
    count = total = 0
    for i in range(n, len(fm_c)):
        n2f, n1f = float(fm_c.iloc[i]), float(fm_c.iloc[i - n])
        n2b, n1b = float(bm_c.iloc[i]), float(bm_c.iloc[i - n])
        if n1f == 0 or n1b == 0: continue
        actual_days = (fm_c.index[i] - fm_c.index[i - n]).days
        if actual_days <= 0: continue
        total += 1
        fund_cagr = (n2f / n1f) ** (365 / actual_days) - 1
        bm_cagr = (n2b / n1b) ** (365 / actual_days) - 1
        if (fund_cagr - bm_cagr) >= threshold: count += 1
    return round(count / total * 100, 4) if total else None

def build_category_avg_series(db, category_id):
    funds = db.execute("SELECT id FROM funds WHERE category_id=? AND type='fund' AND is_active=1", (category_id,)).fetchall()
    if not funds: return pd.Series(dtype=float)
    all_s = []
    for f in funds:
        s = load_series(db, f['id'])
        if len(s) > 0:
            # Normalize to base 100 so averaging is meaningful
            m = get_month_end_navs(s)
            if len(m) > 0:
                normalized = (m / m.iloc[0]) * 100
                all_s.append(normalized)
    if not all_s: return pd.Series(dtype=float)
    df = pd.concat(all_s, axis=1)
    return df.mean(axis=1).dropna()

def get_benchmark_series(db, category_id, bm_type):
    if bm_type == 'category_avg':
        return build_category_avg_series(db, category_id)
    elif bm_type == 'explicit':
        bm_row = db.execute("SELECT id FROM funds WHERE category_id=? AND type='benchmark' AND is_active=1", (category_id,)).fetchone()
        if not bm_row: return pd.Series(dtype=float)
        return load_series(db, bm_row['id'])
    return pd.Series(dtype=float)

def recalculate_all(db, category_id, as_of):
    cat = db.execute("SELECT * FROM categories WHERE id=?", (category_id,)).fetchone()
    if not cat: return
    CY_YEARS = list(range(pd.Timestamp(as_of).year, 2005, -1))
    ROLL_CUTOFF = pd.Timestamp(as_of) - pd.DateOffset(years=10)
    bm_type = cat['bm_type']
    has_bm = bm_type in ('explicit', 'category_avg')
    bm_s = get_benchmark_series(db, category_id, bm_type) if has_bm else pd.Series(dtype=float)
    funds = db.execute("SELECT id,name,type,qualifies_rolling FROM funds WHERE category_id=? AND is_active=1", (category_id,)).fetchall()
    rows = []
    for fund in funds:
        fid, ftype = fund['id'], fund['type']
        s = load_series(db, fid)
        if len(s) == 0: continue
        first_valid = s.dropna().index[0] if len(s.dropna()) > 0 else None
        new_q = 1 if (first_valid and first_valid <= ROLL_CUTOFF) else 0
        db.execute("UPDATE funds SET qualifies_rolling=? WHERE id=?", (new_q, fid))
        if first_valid:
            db.execute("UPDATE funds SET inception_date=? WHERE id=? AND (inception_date IS NULL OR inception_date='')",
                       (str(first_valid.date()), fid))
        si_ret, _ = calc_since_inc(s, as_of)
        si_bm = calc_since_inc_bm(s, bm_s, as_of) if has_bm and ftype == 'fund' and len(bm_s) > 0 else None
        can_outperf = has_bm and new_q and ftype == 'fund' and len(bm_s) > 0
        r = {
            'fund_id': fid, 'as_of_date': as_of,
            'ytd': calc_ytd(s, as_of),
            'ret_1m': calc_trailing(s, as_of, months=1), 'ret_3m': calc_trailing(s, as_of, months=3),
            'ret_6m': calc_trailing(s, as_of, months=6), 'ret_1y': calc_trailing(s, as_of, years=1),
            'ret_2y': calc_trailing(s, as_of, years=2), 'ret_3y': calc_trailing(s, as_of, years=3),
            'ret_5y': calc_trailing(s, as_of, years=5), 'ret_7y': calc_trailing(s, as_of, years=7),
            'ret_10y': calc_trailing(s, as_of, years=10),
            'since_inception': si_ret, 'since_inception_bm': si_bm,
            'roll_3y': calc_rolling(s, 3, as_of) if new_q else None,
            'roll_5y': calc_rolling(s, 5, as_of) if new_q else None,
            'outperf_3y': calc_outperf(s, bm_s, 3, as_of) if can_outperf else None,
            'outperf_5y': calc_outperf(s, bm_s, 5, as_of) if can_outperf else None,
        }
        for yr in range(2006, 2028):
            r[f'cy{yr}'] = calc_cy(s, yr, as_of) if yr in CY_YEARS else None
        rows.append(r)
    if not rows: return
    fids = [r['fund_id'] for r in rows]
    ph = ','.join('?' * len(fids))
    db.execute(f"DELETE FROM precalc_returns WHERE fund_id IN ({ph}) AND as_of_date=?", fids + [as_of])
    cols = 'fund_id,as_of_date,ytd,ret_1m,ret_3m,ret_6m,ret_1y,ret_2y,ret_3y,ret_5y,ret_7y,ret_10y,since_inception,since_inception_bm,roll_3y,roll_5y,outperf_3y,outperf_5y'
    for yr in range(2027, 2005, -1): cols += f',cy{yr}'
    params = ','.join(f':{c}' for c in cols.split(','))
    db.executemany(f"INSERT INTO precalc_returns ({cols}) VALUES ({params})", rows)
    db.commit()

def get_latest_as_of(db, category_id):
    row = db.execute("SELECT MAX(n.nav_date) as d FROM navs n JOIN funds f ON n.fund_id=f.id WHERE f.category_id=? AND f.is_active=1", (category_id,)).fetchone()
    return row['d'] if row and row['d'] else None

# ── FUZZY MATCHING ─────────────────────────────────────────────────────────────
def fuzzy_match(raw_name, all_funds):
    u = raw_name.lower().strip()
    u_clean = u
    for pfx in ['bm:', 'si:']:
        if u_clean.startswith(pfx): u_clean = u_clean[len(pfx):].strip()
    for f in all_funds:
        fn = f['name'].lower()
        if fn == u or fn == u_clean: return f['id']
    for f in all_funds:
        fn = f['name'].lower()
        if u_clean in fn or fn in u_clean: return f['id']
    words = [w for w in u_clean.split() if len(w) > 2]
    best_score, best_id = 0, None
    for f in all_funds:
        fn = f['name'].lower()
        score = sum(len(w) for w in words if w in fn)
        if score > best_score: best_score, best_id = score, f['id']
    return best_id if best_score > 5 else None

def parse_prefix(col_name):
    clean = str(col_name).strip()
    if not clean or clean.lower() == 'nan': return None, None
    upper = clean.upper().replace(' ', '')
    if upper.startswith('BM:') and ':' in clean: return clean[clean.index(':') + 1:].strip(), 'benchmark'
    elif upper.startswith('SI:') and ':' in clean: return clean[clean.index(':') + 1:].strip(), 'strategy'
    return clean, 'fund'

# ── NAV UPLOAD (FIRST TIME) ───────────────────────────────────────────────────
@app.route('/api/upload_navs', methods=['POST'])
def api_upload_navs():
    pwd = request.form.get('password', '')
    if not check_admin_pwd(pwd): return jsonify({'success': False, 'error': 'Incorrect password'}), 403
    file = request.files.get('file')
    category_code = request.form.get('category', '')
    if not file or file.filename == '': return jsonify({'success': False, 'error': 'No file'}), 400
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category_code,)).fetchone()
    if not cat: return jsonify({'success': False, 'error': 'Invalid category'}), 400
    try:
        buf = io.BytesIO(file.read()); fname = file.filename.lower()
        if fname.endswith('.csv'): df = pd.read_csv(buf)
        else:
            raw = pd.read_excel(buf, header=None, engine='openpyxl')
            hdr_row = 0
            for i in range(min(8, len(raw))):
                if 'date' in ' '.join(str(v).lower() for v in raw.iloc[i]): hdr_row = i; break
            buf.seek(0); df = pd.read_excel(buf, header=hdr_row, engine='openpyxl')
        date_col = next((c for c in df.columns if 'date' in str(c).lower()), df.columns[0])
        df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
        df = df.dropna(subset=[date_col]).sort_values(date_col)
        fund_columns = [c for c in df.columns if c != date_col]
        if not fund_columns: return jsonify({'success': False, 'error': 'No fund columns'}), 400
        existing = db.execute("SELECT id,name,type FROM funds WHERE category_id=?", (cat['id'],)).fetchall()
        existing_map = {f['name'].lower(): f for f in existing}
        created, matched, skipped = [], [], []
        col_map = {}
        for col in fund_columns:
            dn, ft = parse_prefix(col)
            if not dn: continue
            valid = df[col].apply(lambda v: pd.notna(v) and str(v).strip() not in ['', '0', 'nan']).sum()
            if valid == 0: skipped.append(str(col)); continue
            found = existing_map.get(dn.lower())
            if found: col_map[col] = found['id']; matched.append(dn)
            else:
                fid = fuzzy_match(dn, existing)
                if fid: col_map[col] = fid; matched.append(dn)
                else:
                    short = dn.replace(' Fund-Reg(G)', '').replace(' Fund(G)', '')
                    db.execute("INSERT INTO funds(category_id,name,short_name,type,is_active) VALUES(?,?,?,?,1)",
                               (cat['id'], dn, short, ft)); db.commit()
                    nf = db.execute("SELECT id FROM funds WHERE name=? AND category_id=?", (dn, cat['id'])).fetchone()
                    if nf:
                        col_map[col] = nf['id']
                        db.execute("INSERT OR IGNORE INTO fund_details(fund_id) VALUES(?)", (nf['id'],))
                        created.append(dn)
        nav_rows = []
        for _, row in df.iterrows():
            nd = str(row[date_col].date())
            for cn, fid in col_map.items():
                rv = row[cn]
                if pd.isna(rv): continue
                try:
                    val = float(rv)
                    if val > 0: nav_rows.append((fid, nd, round(val, 4)))
                except: continue
        if nav_rows:
            db.executemany("INSERT OR REPLACE INTO navs(fund_id,nav_date,nav_value) VALUES(?,?,?)", nav_rows)
            db.commit()
        fund_nav_counts = {}
        for fid in set(r[0] for r in nav_rows):
            cnt = sum(1 for r in nav_rows if r[0] == fid)
            nm = next((k for k, v in col_map.items() if v == fid), str(fid))
            dn2, _ = parse_prefix(nm)
            fund_nav_counts[dn2 or nm] = cnt
        latest = get_latest_as_of(db, cat['id'])
        if latest: recalculate_all(db, cat['id'], latest)
        return jsonify({'success': True, 'created_funds': created, 'matched_funds': matched,
                        'skipped_cols': skipped, 'total_navs': len(nav_rows), 'dates_count': len(df),
                        'fund_nav_counts': fund_nav_counts, 'latest_as_of': latest,
                        'message': f"Processed {len(fund_columns)} columns. Created {len(created)}, matched {len(matched)}. {len(nav_rows)} NAVs across {len(df)} dates."})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ── NAV UPDATE (ONGOING) ──────────────────────────────────────────────────────
@app.route('/api/upload_update', methods=['POST'])
def api_upload_update():
    pwd = request.form.get('password', '')
    if not check_admin_pwd(pwd): return jsonify({'success': False, 'error': 'Incorrect password'}), 403
    files = request.files.getlist('files')
    category_code = request.form.get('category', '')
    if not files or all(f.filename == '' for f in files): return jsonify({'success': False, 'error': 'No files'}), 400
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category_code,)).fetchone()
    if not cat: return jsonify({'success': False, 'error': 'Invalid category'}), 400
    all_funds = db.execute("SELECT id,name FROM funds WHERE category_id=? AND is_active=1", (cat['id'],)).fetchall()
    total_rows, file_results, parse_errors = 0, [], []
    for file in files:
        if file.filename == '': continue
        buf = io.BytesIO(file.read())
        try:
            fn = file.filename.lower()
            if fn.endswith('.csv'): dff = pd.read_csv(buf)
            else:
                raw = pd.read_excel(buf, header=None, engine='openpyxl')
                hr = 0
                for i in range(min(6, len(raw))):
                    rs = ' '.join(str(v).lower() for v in raw.iloc[i])
                    if any(k in rs for k in ['fund', 'name', 'index', 'scheme', 'date']): hr = i; break
                buf.seek(0); dff = pd.read_excel(buf, header=hr, engine='openpyxl')
            has_date = any('date' in str(c).lower() for c in dff.columns)
            if has_date and len(dff.columns) > 3:
                dc = next(c for c in dff.columns if 'date' in str(c).lower())
                dff[dc] = pd.to_datetime(dff[dc], dayfirst=True, errors='coerce')
                dff = dff.dropna(subset=[dc])
                fcs = [c for c in dff.columns if c != dc]
                mt, nd = 0, []
                for col in fcs:
                    dn, _ = parse_prefix(col)
                    if not dn: continue
                    fid = None
                    for f in all_funds:
                        if f['name'].lower() == dn.lower(): fid = f['id']; break
                    if not fid: fid = fuzzy_match(dn, all_funds)
                    if fid:
                        mt += 1
                        for _, row in dff.iterrows():
                            rv = row[col]
                            if pd.isna(rv): continue
                            try:
                                val = float(rv)
                                if val > 0: nd.append((fid, str(row[dc].date()), round(val, 4)))
                            except: continue
                if nd:
                    db.executemany("INSERT OR REPLACE INTO navs(fund_id,nav_date,nav_value) VALUES(?,?,?)", nd)
                    total_rows += len(nd)
                file_results.append({'filename': file.filename, 'matched': mt, 'navs': len(nd)})
            else:
                cols = [str(c).lower() for c in dff.columns]
                nc = next((dff.columns[i] for i, c in enumerate(cols) if any(k in c for k in ['fund', 'name', 'index', 'scheme'])), dff.columns[0])
                vc = next((dff.columns[i] for i, c in enumerate(cols) if any(k in c for k in ['nav', 'value', 'close', 'price', 'level'])), dff.columns[1])
                dcn = next((dff.columns[i] for i, c in enumerate(cols) if any(k in c for k in ['date', 'as of'])), None)
                nav_date = None
                if dcn:
                    vals = dff[dcn].dropna()
                    if len(vals) > 0:
                        try: nav_date = str(pd.to_datetime(vals.iloc[0], dayfirst=True).date())
                        except: pass
                if not nav_date:
                    m = re.search(r'(\d{1,2})[-._](\d{1,2})[-._](\d{2,4})', file.filename)
                    if m:
                        y = m.group(3)
                        if len(y) == 2: y = '20' + y
                        try: nav_date = str(pd.to_datetime(f"{y}-{m.group(2)}-{m.group(1)}").date())
                        except: pass
                if not nav_date: parse_errors.append(f"{file.filename}: No date detected"); continue
                mt, nd = 0, []
                for _, row in dff.iterrows():
                    rn = str(row[nc]).strip()
                    if not rn or rn.lower() == 'nan': continue
                    try: nv = float(row[vc])
                    except: continue
                    if pd.isna(nv) or nv <= 0: continue
                    fid = fuzzy_match(rn, all_funds)
                    if fid: nd.append((fid, nav_date, round(nv, 4))); mt += 1
                if nd:
                    db.executemany("INSERT OR REPLACE INTO navs(fund_id,nav_date,nav_value) VALUES(?,?,?)", nd)
                    total_rows += len(nd)
                file_results.append({'filename': file.filename, 'nav_date': nav_date, 'matched': mt, 'navs': len(nd)})
        except Exception as e:
            parse_errors.append(f"{file.filename}: {str(e)}")
    db.commit()
    latest = get_latest_as_of(db, cat['id'])
    if latest: recalculate_all(db, cat['id'], latest)
    return jsonify({'success': True, 'files': file_results, 'total_navs': total_rows,
                    'latest_as_of': latest, 'parse_errors': parse_errors,
                    'message': f"Processed {len(file_results)} file(s). {total_rows} NAVs stored."})

# ── FUND DETAILS UPLOAD ───────────────────────────────────────────────────────
@app.route('/api/update_fund_details', methods=['POST'])
def api_update_fund_details():
    pwd = request.form.get('password', '')
    if not check_admin_pwd(pwd): return jsonify({'success': False, 'error': 'Incorrect password'}), 403
    file = request.files.get('file')
    category_code = request.form.get('category', '')
    quarter = request.form.get('quarter', '').strip()
    if not file or file.filename == '': return jsonify({'success': False, 'error': 'No file'}), 400
    # ── Read file: support both CSV and Excel ──────────────────────────────────
    try:
        buf = io.BytesIO(file.read())
        fname = file.filename.lower()
        if fname.endswith('.csv'):
            df = pd.read_csv(buf)
        else:
            df = pd.read_excel(buf, engine='openpyxl')
    except Exception as read_err:
        return jsonify({'success': False, 'error': f'Could not read file: {str(read_err)}'}), 400
    # Strip whitespace from all column names
    df.columns = [str(c).strip() for c in df.columns]
    try:
        db = get_db()
        cat = db.execute("SELECT id FROM categories WHERE code=?", (category_code,)).fetchone()
        cid = cat['id'] if cat else None
        def safe(v):
            try:
                if pd.isna(v): return None
            except (TypeError, ValueError): pass
            if str(v).strip() in ('', '-', 'nan', 'None'): return None
            try: return round(float(str(v).replace(',', '')), 4)
            except: return None
        def safestr(v):
            if v is None: return None
            try:
                if pd.isna(v): return None
            except (TypeError, ValueError): pass
            s = str(v).strip()
            return None if s in ('', 'nan', 'None', '-') else s
        # Exact-match-first column finder
        def col(name, exact=False):
            name_l = name.lower().strip()
            for c in df.columns:
                if str(c).lower().strip() == name_l: return c
            if exact: return None
            for c in df.columns:
                if name_l in str(c).lower(): return c
            return None
        nc    = col('fund name') or col('name') or df.columns[0]
        bc    = col('beta')
        ec    = col('expense ratio', exact=True) or col('exp ratio', exact=True) or col('expense') or col('exp')
        sc    = col('std dev', exact=True) or col('std_dev', exact=True) or col('std dev') or col('std.')
        ic    = col('info ratio', exact=True) or col('info_ratio', exact=True) or col('info ratio') or col('information')
        m1    = col('manager 1', exact=True) or col('manager1', exact=True) or col('manager 1')
        m2    = col('manager 2', exact=True) or col('manager2', exact=True) or col('manager 2')
        ac    = col('aum', exact=True) or col('aum')
        lcc   = col('lc %', exact=True) or col('lc%', exact=True) or col('lc %') or col('large cap')
        mcc   = col('mc %', exact=True) or col('mc%', exact=True) or col('mc %') or col('mid cap')
        scc   = col('sc %', exact=True) or col('sc%', exact=True) or col('sc %') or col('small cap')
        cashc = col('cash %', exact=True) or col('cash%', exact=True) or col('cash')
        s1n   = col('sector 1', exact=True) or col('sector1', exact=True)
        s1p   = col('sector 1 %', exact=True) or col('sector1 %', exact=True) or col('sector1%', exact=True)
        s2n   = col('sector 2', exact=True) or col('sector2', exact=True)
        s2p   = col('sector 2 %', exact=True) or col('sector2 %', exact=True) or col('sector2%', exact=True)
        s3n   = col('sector 3', exact=True) or col('sector3', exact=True)
        s3p   = col('sector 3 %', exact=True) or col('sector3 %', exact=True) or col('sector3%', exact=True)
        eqc   = col('equity %', exact=True) or col('equity%', exact=True) or col('equity %') or col('equity')
        neqc  = col('net equity %', exact=True) or col('net equity', exact=True) or col('net_equity', exact=True) or col('net equity')
        dbc   = col('debt %', exact=True) or col('debt%', exact=True) or col('debt %') or col('debt')
        otc   = col('others %', exact=True) or col('others%', exact=True) or col('others %') or col('others')
        gldc  = col('gold %', exact=True) or col('gold%', exact=True) or col('gold %') or col('gold')
        slvc  = col('silver %', exact=True) or col('silver%', exact=True) or col('silver %') or col('silver')
        rtc   = col('reits %') or col('reits') or col('reit')
        incc  = col('inception')
        afs = db.execute(
            "SELECT id,name FROM funds WHERE category_id=? AND is_active=1", (cid,)
        ).fetchall() if cid else db.execute(
            "SELECT id,name FROM funds WHERE is_active=1"
        ).fetchall()

        now_ts = datetime.now().isoformat()
        updated, unmatched = 0, []
        # Collect all rows first, then batch-write in a single transaction
        detail_upserts = []   # (fund_id, field_dict)
        history_rows   = []   # param dicts for INSERT OR REPLACE
        inception_upd  = []   # (date_str, fund_id)

        for _, row in df.iterrows():
            rn = safestr(row[nc])
            if not rn: continue
            fid = fuzzy_match(rn, afs)
            if not fid: unmatched.append(rn); continue
            u = {}
            if bc:    u['beta']          = safe(row[bc])
            if ec:    u['expense_ratio'] = safe(row[ec])
            if sc:    u['std_dev']       = safe(row[sc])
            if ic:    u['info_ratio']    = safe(row[ic])
            if m1:    u['manager1']      = safestr(row[m1])
            if m2:    u['manager2']      = safestr(row[m2])
            if ac:    u['aum_latest']    = safe(row[ac])
            if lcc:   u['lc_pct']        = safe(row[lcc])
            if mcc:   u['mc_pct']        = safe(row[mcc])
            if scc:   u['sc_pct']        = safe(row[scc])
            if cashc: u['cash_pct']      = safe(row[cashc])
            if s1n:   u['sector1_name']  = safestr(row[s1n])
            if s1p:   u['sector1_pct']   = safe(row[s1p])
            if s2n:   u['sector2_name']  = safestr(row[s2n])
            if s2p:   u['sector2_pct']   = safe(row[s2p])
            if s3n:   u['sector3_name']  = safestr(row[s3n])
            if s3p:   u['sector3_pct']   = safe(row[s3p])
            if eqc:   u['equity_pct']    = safe(row[eqc])
            if neqc:  u['net_equity_pct']= safe(row[neqc])
            if dbc:   u['debt_pct']      = safe(row[dbc])
            if otc:   u['others_pct']    = safe(row[otc])
            if gldc:  u['gold_pct']      = safe(row[gldc])
            if slvc:  u['silver_pct']    = safe(row[slvc])
            if rtc:   u['reits_pct']     = safe(row[rtc])
            u['updated_at'] = now_ts
            if incc:
                iv = safestr(row[incc])
                if iv:
                    try: inception_upd.append((str(pd.to_datetime(iv, dayfirst=True).date()), fid))
                    except: pass
            if u:
                detail_upserts.append((fid, u))
                if quarter:
                    h = {'fund_id': fid, 'quarter': quarter}
                    for k in ['aum_latest','lc_pct','mc_pct','sc_pct',
                              'sector1_name','sector1_pct','sector2_name','sector2_pct',
                              'sector3_name','sector3_pct',
                              'equity_pct','net_equity_pct','debt_pct','others_pct',
                              'gold_pct','silver_pct','reits_pct']:
                        h['aum' if k == 'aum_latest' else k] = u.get(k)
                    history_rows.append(h)
                updated += 1

        # ── Single transaction: all writes at once ─────────────────────────────
        with db:
            for fid, u in detail_upserts:
                db.execute("INSERT OR IGNORE INTO fund_details(fund_id) VALUES(?)", (fid,))
                db.execute(
                    f"UPDATE fund_details SET {', '.join(f'{k}=?' for k in u)} WHERE fund_id=?",
                    list(u.values()) + [fid]
                )
            for h in history_rows:
                hcols   = ','.join(h.keys())
                hparams = ','.join(f':{k}' for k in h.keys())
                db.execute(f"INSERT OR REPLACE INTO fund_details_history ({hcols}) VALUES ({hparams})", h)
            for inc_date, fid in inception_upd:
                db.execute("UPDATE funds SET inception_date=? WHERE id=?", (inc_date, fid))

        msg = f'Updated details for {updated} funds'
        if unmatched: msg += f'. {len(unmatched)} unmatched: {", ".join(unmatched[:10])}'
        return jsonify({'success': True, 'updated': updated, 'unmatched': unmatched[:10], 'message': msg})
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'detail': traceback.format_exc()}), 500

# ── EXCEL DOWNLOAD ─────────────────────────────────────────────────────────────
def build_excel(db, category_id, as_of):
    wb = openpyxl.Workbook()
    thin = Side(style='thin', color='CCCCCC'); bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    hfill = PatternFill('solid', start_color='1E3A5F')
    bmfill = PatternFill('solid', start_color='FEF3C7')
    alt = PatternFill('solid', start_color='F0F4F8')
    ctr = Alignment(horizontal='center', vertical='center')
    def hc(ws, r, c, v, w=None):
        cl = ws.cell(r, c, v); cl.font = hf; cl.fill = hfill; cl.alignment = ctr; cl.border = bdr
        if w: ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    def dc(ws, r, c, v, fmt=None, fill=None):
        cl = ws.cell(r, c, v); cl.font = Font(name='Arial', size=9)
        cl.alignment = Alignment(horizontal='right' if isinstance(v, (int, float)) else 'left', vertical='center')
        cl.border = bdr
        if fmt: cl.number_format = fmt
        if fill: cl.fill = fill
    cat = db.execute("SELECT * FROM categories WHERE id=?", (category_id,)).fetchone()
    has_bm = cat['bm_type'] in ('explicit', 'category_avg') if cat else False
    funds = db.execute("""
        SELECT f.id,f.name,f.type,
            pr.ytd,pr.ret_1m,pr.ret_3m,pr.ret_6m,pr.ret_1y,pr.ret_2y,pr.ret_3y,pr.ret_5y,pr.ret_7y,pr.ret_10y,
            pr.since_inception,pr.since_inception_bm,pr.roll_3y,pr.roll_5y,pr.outperf_3y,pr.outperf_5y,
            pr.cy2027,pr.cy2026,pr.cy2025,pr.cy2024,pr.cy2023,pr.cy2022,pr.cy2021,pr.cy2020,
            pr.cy2019,pr.cy2018,pr.cy2017,pr.cy2016,pr.cy2015,
            fd.aum_latest,fd.expense_ratio,fd.beta,fd.info_ratio,fd.std_dev,
            fd.lc_pct,fd.mc_pct,fd.sc_pct,f.inception_date
        FROM funds f LEFT JOIN precalc_returns pr ON pr.fund_id=f.id AND pr.as_of_date=?
        LEFT JOIN fund_details fd ON fd.fund_id=f.id
        WHERE f.category_id=? AND f.is_active=1
        ORDER BY f.type DESC, fd.aum_latest DESC NULLS LAST
    """, (as_of, category_id)).fetchall()
    cy = pd.Timestamp(as_of).year; CY = list(range(cy, cy - 10, -1))
    # Sheet 1: Trailing
    TK = ['ytd', 'ret_1m', 'ret_3m', 'ret_6m', 'ret_1y', 'ret_2y', 'ret_3y', 'ret_5y', 'ret_7y', 'ret_10y', 'since_inception']
    TH = ['YTD', '1M', '3M', '6M', '1Y', '2Y', '3Y', '5Y', '7Y', '10Y', 'Since Inc.']
    if has_bm: TK.append('since_inception_bm'); TH.append('SI BM')
    ws1 = wb.active; ws1.title = 'Trailing Returns'
    ws1.cell(1, 1, f'Trailing Returns — As of {as_of}').font = Font(bold=True, name='Arial', size=10)
    hc(ws1, 2, 1, 'Fund / Index', 40)
    for i, (h, w) in enumerate(zip(TH, [9] * len(TH)), 2): hc(ws1, 2, i, h, w)
    for ri, f in enumerate(funds, 3):
        fl = bmfill if f['type'] == 'benchmark' else (alt if ri % 2 == 0 else None)
        dc(ws1, ri, 1, f['name'], fill=fl)
        for ci, k in enumerate(TK, 2):
            v = f[k]; dc(ws1, ri, ci, v / 100 if v is not None else None, '0.00%', fl)
    # Sheet 2: Calendar Year
    ws2 = wb.create_sheet('Calendar Year')
    ws2.cell(1, 1, f'Calendar Year — As of {as_of}').font = Font(bold=True, name='Arial', size=10)
    hc(ws2, 2, 1, 'Fund / Index', 40)
    for i, yr in enumerate(CY, 2): hc(ws2, 2, i, str(yr), 9)
    for ri, f in enumerate(funds, 3):
        fl = bmfill if f['type'] == 'benchmark' else (alt if ri % 2 == 0 else None)
        dc(ws2, ri, 1, f['name'], fill=fl)
        for ci, yr in enumerate(CY, 2):
            try: v = f[f'cy{yr}']
            except: v = None
            dc(ws2, ri, ci, v / 100 if v is not None else None, '0.00%', fl)
    # Sheet 3: Rolling
    ws3 = wb.create_sheet('Rolling Returns')
    ws3.cell(1, 1, f'Rolling Returns — As of {as_of}').font = Font(bold=True, name='Arial', size=10)
    RH = ['3Y Rolling', '5Y Rolling']
    if has_bm: RH += ['% Outperf (3Y)', '% Outperf (5Y)']
    RH += ['Since Inc.']
    if has_bm: RH.append('SI BM')
    RH += ['Beta', 'Info Ratio', 'Std Dev']
    hc(ws3, 2, 1, 'Fund / Index', 40)
    for i, (h, w) in enumerate(zip(RH, [14] * len(RH)), 2): hc(ws3, 2, i, h, w)
    for ri, f in enumerate(funds, 3):
        fl = bmfill if f['type'] == 'benchmark' else (alt if ri % 2 == 0 else None)
        dc(ws3, ri, 1, f['name'], fill=fl)
        ci = 2
        dc(ws3, ri, ci, f['roll_3y'] / 100 if f['roll_3y'] is not None else None, '0.00%', fl); ci += 1
        dc(ws3, ri, ci, f['roll_5y'] / 100 if f['roll_5y'] is not None else None, '0.00%', fl); ci += 1
        if has_bm:
            dc(ws3, ri, ci, f['outperf_3y'] / 100 if f['outperf_3y'] is not None else None, '0.00%', fl); ci += 1
            dc(ws3, ri, ci, f['outperf_5y'] / 100 if f['outperf_5y'] is not None else None, '0.00%', fl); ci += 1
        dc(ws3, ri, ci, f['since_inception'] / 100 if f['since_inception'] is not None else None, '0.00%', fl); ci += 1
        if has_bm:
            dc(ws3, ri, ci, f['since_inception_bm'] / 100 if f['since_inception_bm'] is not None else None, '0.00%', fl); ci += 1
        dc(ws3, ri, ci, f['beta'], '0.0000', fl); ci += 1
        dc(ws3, ri, ci, f['info_ratio'], '0.0000', fl); ci += 1
        dc(ws3, ri, ci, f['std_dev'], '0.00', fl)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ── ROUTES ─────────────────────────────────────────────────────────────────────
@app.route('/')
@dash_login_required
def index(): return redirect(url_for('dashboard', category='mf_largecap'))

@app.route('/dashboard/<category>')
@dash_login_required
def dashboard(category):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category,)).fetchone()
    if not cat: return redirect(url_for('dashboard', category='mf_largecap'))
    cats = db.execute("SELECT * FROM categories WHERE is_active=1 ORDER BY display_order").fetchall()
    as_of = get_latest_as_of(db, cat['id'])
    ns = {'mf': {'equity': [], 'hybrids': []}, 'nmf': {'non_mf': []}, 'idx': {'indices': []}}
    for c in cats:
        u, pg = c['universe'], c['parent_group']
        if u in ns and pg in ns[u]: ns[u][pg].append(dict(c))
    return render_template('dashboard.html', category=cat, categories=cats, as_of=as_of or '', nav_structure=ns)

@app.route('/api/categories')
def api_categories():
    db = get_db()
    cats = db.execute("SELECT * FROM categories WHERE is_active=1 ORDER BY display_order").fetchall()
    result = {}
    for c in cats:
        cd = dict(c)
        cd['fund_count'] = db.execute("SELECT COUNT(*) as cnt FROM funds WHERE category_id=? AND is_active=1",
                                       (c['id'],)).fetchone()['cnt']
        u, pg = c['universe'], c['parent_group']
        result.setdefault(u, {}).setdefault(pg, []).append(cd)
    return jsonify(result)

@app.route('/api/funds/<category>')
def api_funds(category):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    as_of = request.args.get('as_of') or get_latest_as_of(db, cat['id'])
    if not as_of:
        return jsonify({'funds': [], 'as_of': None, 'latest_as_of': None,
                        'bm_type': cat['bm_type'], 'empty': True})
    funds = db.execute("""
        SELECT f.id,f.name,f.short_name,f.type,f.qualifies_rolling,f.inception_date,
            pr.ytd,pr.ret_1m,pr.ret_3m,pr.ret_6m,pr.ret_1y,pr.ret_2y,pr.ret_3y,
            pr.ret_5y,pr.ret_7y,pr.ret_10y,pr.since_inception,pr.since_inception_bm,
            pr.roll_3y,pr.roll_5y,pr.outperf_3y,pr.outperf_5y,
            pr.cy2027,pr.cy2026,pr.cy2025,pr.cy2024,pr.cy2023,pr.cy2022,pr.cy2021,pr.cy2020,
            pr.cy2019,pr.cy2018,pr.cy2017,pr.cy2016,pr.cy2015,pr.cy2014,pr.cy2013,
            pr.cy2012,pr.cy2011,pr.cy2010,pr.cy2009,pr.cy2008,pr.cy2007,pr.cy2006,
            fd.manager1,fd.manager2,fd.aum_latest,fd.expense_ratio,fd.std_dev,
            fd.sharpe,fd.beta,fd.info_ratio,fd.volatility,
            fd.lc_pct,fd.mc_pct,fd.sc_pct,fd.cash_pct,
            fd.sector1_name,fd.sector1_pct,fd.sector2_name,fd.sector2_pct,fd.sector3_name,fd.sector3_pct,
            fd.equity_pct,fd.net_equity_pct,fd.debt_pct,fd.others_pct,fd.gold_pct,fd.silver_pct,fd.reits_pct
        FROM funds f LEFT JOIN precalc_returns pr ON pr.fund_id=f.id AND pr.as_of_date=?
        LEFT JOIN fund_details fd ON fd.fund_id=f.id
        WHERE f.category_id=? AND f.is_active=1
        ORDER BY f.type DESC, fd.aum_latest DESC NULLS LAST
    """, (as_of, cat['id'])).fetchall()
    if funds and all(f['ret_1y'] is None for f in funds):
        recalculate_all(db, cat['id'], as_of)
        return api_funds(category)
    return jsonify({'funds': [dict(f) for f in funds], 'as_of': as_of,
                    'latest_as_of': get_latest_as_of(db, cat['id']), 'bm_type': cat['bm_type']})

@app.route('/api/funds_as_of/<category>')
def api_funds_as_of(category):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    as_of = request.args.get('as_of')
    if not as_of: return jsonify({'error': 'as_of required'}), 400
    # Always recalculate for date changes (rolling returns depend on as_of)
    recalculate_all(db, cat['id'], as_of)
    return api_funds(category)

@app.route('/api/custom_return', methods=['POST'])
def api_custom_return():
    data = request.get_json()
    cat_code, fd, td = data.get('category'), data.get('from_date'), data.get('to_date')
    fids = data.get('fund_ids', [])
    if not all([cat_code, fd, td]): return jsonify({'error': 'Missing params'}), 400
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (cat_code,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    if fids:
        ph = ','.join('?' * len(fids))
        funds = db.execute(f"SELECT id,name,short_name,type FROM funds WHERE id IN ({ph}) AND is_active=1", fids).fetchall()
    else:
        funds = db.execute("SELECT id,name,short_name,type FROM funds WHERE category_id=? AND is_active=1",
                           (cat['id'],)).fetchall()
    results = []
    for f in funds:
        s = load_series(db, f['id'])
        ar, anr, days = calc_custom(s, fd, td)
        if ar is not None:
            results.append({'fund_id': f['id'], 'name': f['name'], 'short_name': f['short_name'],
                            'type': f['type'], 'abs_return': ar, 'ann_return': anr, 'days': days,
                            'nav_from': get_nav_at(s, fd), 'nav_to': get_nav_at(s, td)})
    results.sort(key=lambda x: x['abs_return'], reverse=True)
    return jsonify({'results': results, 'from_date': fd, 'to_date': td})

@app.route('/api/outperf_custom', methods=['POST'])
def api_outperf_custom():
    data = request.get_json()
    cat_code, thr = data.get('category'), data.get('threshold', 0.5)
    if not cat_code: return jsonify({'error': 'Missing category'}), 400
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (cat_code,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    as_of = data.get('as_of') or get_latest_as_of(db, cat['id'])
    if not as_of: return jsonify({'error': 'No data'}), 400
    bm_s = get_benchmark_series(db, cat['id'], cat['bm_type'])
    if len(bm_s) == 0: return jsonify({'error': 'No BM'}), 400
    funds = db.execute("SELECT id,name FROM funds WHERE category_id=? AND type='fund' AND is_active=1 AND qualifies_rolling=1",
                       (cat['id'],)).fetchall()
    tv = float(thr) / 100; results = []
    for f in funds:
        s = load_series(db, f['id'])
        results.append({'fund_id': f['id'], 'name': f['name'],
                        'outperf_3y': calc_outperf(s, bm_s, 3, as_of, threshold=tv),
                        'outperf_5y': calc_outperf(s, bm_s, 5, as_of, threshold=tv)})
    return jsonify({'results': results, 'threshold': thr})

@app.route('/api/download/<category>')
def api_download(category):
    db = get_db()
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    as_of = request.args.get('as_of') or get_latest_as_of(db, cat['id'])
    if not as_of: return jsonify({'error': 'No data'}), 400
    buf = build_excel(db, cat['id'], as_of)
    return send_file(buf, as_attachment=True,
                     download_name=f"Daksham_{cat['name'].replace(' ', '_')}_{as_of}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/fund_history/<int:fund_id>')
def api_fund_history(fund_id):
    return jsonify({'history': [dict(r) for r in get_db().execute(
        "SELECT * FROM fund_details_history WHERE fund_id=? ORDER BY quarter DESC", (fund_id,)).fetchall()]})

@app.route('/api/fund_details_view')
def api_fund_details_view():
    if not session.get('admin_logged_in'): return jsonify({'error': 'Unauthorized'}), 403
    db = get_db()
    category = request.args.get('category', '')
    quarter = request.args.get('quarter', '')
    if category:
        cat = db.execute("SELECT id FROM categories WHERE code=?", (category,)).fetchone()
        if not cat: return jsonify({'error': 'Not found'}), 404
        funds = db.execute("""
            SELECT f.id,f.name,f.type,f.inception_date,fd.*
            FROM funds f LEFT JOIN fund_details fd ON fd.fund_id=f.id
            WHERE f.category_id=? AND f.is_active=1
            ORDER BY f.type DESC, fd.aum_latest DESC NULLS LAST
        """, (cat['id'],)).fetchall()
    else:
        funds = db.execute("""
            SELECT f.id,f.name,f.type,f.inception_date,fd.*
            FROM funds f LEFT JOIN fund_details fd ON fd.fund_id=f.id
            WHERE f.is_active=1 ORDER BY f.type DESC, fd.aum_latest DESC NULLS LAST
        """).fetchall()
    quarters = db.execute("SELECT DISTINCT quarter FROM fund_details_history ORDER BY quarter DESC").fetchall()
    hist_data = {}
    if quarter:
        rows = db.execute("SELECT * FROM fund_details_history WHERE quarter=?", (quarter,)).fetchall()
        for r in rows: hist_data[r['fund_id']] = dict(r)
    return jsonify({'funds': [dict(f) for f in funds],
                    'quarters': [q['quarter'] for q in quarters], 'hist_data': hist_data})

# ── NAV HISTORY ────────────────────────────────────────────────────────────────
@app.route('/api/nav_history')
def api_nav_history():
    if not session.get('admin_logged_in'): return jsonify({'error': 'Unauthorized'}), 403
    db = get_db(); category = request.args.get('category', 'mf_largecap')
    from_d, to_d = request.args.get('from_date'), request.args.get('to_date')
    page, per_page = int(request.args.get('page', 1)), 50
    fids_filter = request.args.getlist('fund_ids')
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    if not to_d: to_d = get_latest_as_of(db, cat['id']) or str(date.today())
    if not from_d: from_d = str((pd.Timestamp(to_d) - pd.DateOffset(years=2)).date())
    funds = db.execute("SELECT id,name,short_name,type FROM funds WHERE category_id=? AND is_active=1 ORDER BY type DESC, name",
                       (cat['id'],)).fetchall()
    if fids_filter:
        fids_filter = [int(x) for x in fids_filter]
        funds = [f for f in funds if f['id'] in fids_filter]
    fund_ids = [f['id'] for f in funds]
    if fund_ids:
        ph = ','.join('?' * len(fund_ids))
        actual = db.execute(f"SELECT DISTINCT nav_date FROM navs WHERE fund_id IN ({ph}) AND nav_date BETWEEN ? AND ? ORDER BY nav_date DESC",
                            fund_ids + [from_d, to_d]).fetchall()
    else: actual = []
    dr = [r['nav_date'] for r in actual]
    td = len(dr); tp = max(1, (td + per_page - 1) // per_page)
    pd_dates = dr[(page - 1) * per_page:page * per_page]
    if fund_ids and pd_dates:
        ph = ','.join('?' * len(fund_ids)); dph = ','.join('?' * len(pd_dates))
        navs = db.execute(f"SELECT fund_id,nav_date,nav_value FROM navs WHERE fund_id IN ({ph}) AND nav_date IN ({dph})",
                          fund_ids + pd_dates).fetchall()
    else: navs = []
    nl = {}
    for n in navs: nl.setdefault(n['nav_date'], {})[n['fund_id']] = n['nav_value']
    rows = []
    for d in pd_dates:
        row = {'date': d, 'navs': {}}; dn = nl.get(d, {})
        for f in funds: row['navs'][f['id']] = dn.get(f['id'])
        rows.append(row)
    af = [dict(f) for f in db.execute("SELECT id,name,short_name,type FROM funds WHERE category_id=? AND is_active=1 ORDER BY type DESC,name",
                                       (cat['id'],)).fetchall()]
    return jsonify({'funds': [dict(f) for f in funds], 'rows': rows, 'page': page,
                    'total_pages': tp, 'total_dates': td, 'from_date': from_d, 'to_date': to_d, 'all_funds': af})

@app.route('/api/nav_history_download')
def api_nav_history_download():
    if not session.get('admin_logged_in'): return jsonify({'error': 'Unauthorized'}), 403
    db = get_db(); category = request.args.get('category', 'mf_largecap')
    from_d, to_d = request.args.get('from_date'), request.args.get('to_date')
    cat = db.execute("SELECT * FROM categories WHERE code=?", (category,)).fetchone()
    if not cat: return jsonify({'error': 'Not found'}), 404
    if not to_d: to_d = get_latest_as_of(db, cat['id']) or str(date.today())
    if not from_d: from_d = str((pd.Timestamp(to_d) - pd.DateOffset(years=1)).date())
    funds = db.execute("SELECT id,name FROM funds WHERE category_id=? AND is_active=1 ORDER BY type DESC, name",
                       (cat['id'],)).fetchall()
    if not funds: return jsonify({'error': 'No funds'}), 400
    fids = [f['id'] for f in funds]; ph = ','.join('?' * len(fids))
    navs = db.execute(f"SELECT fund_id,nav_date,nav_value FROM navs WHERE fund_id IN ({ph}) AND nav_date BETWEEN ? AND ? ORDER BY nav_date DESC",
                      fids + [from_d, to_d]).fetchall()
    data = {}
    for n in navs: data.setdefault(n['nav_date'], {})[n['fund_id']] = n['nav_value']
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'NAV History'; ws.cell(1, 1, 'Date')
    for i, f in enumerate(funds, 2): ws.cell(1, i, f['name'])
    for ri, (dt, vals) in enumerate(sorted(data.items(), reverse=True), 2):
        ws.cell(ri, 1, dt)
        for ci, f in enumerate(funds, 2): ws.cell(ri, ci, vals.get(f['id']))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"NAV_History_{cat['name'].replace(' ', '_')}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── BULK DELETE ────────────────────────────────────────────────────────────────
@app.route('/admin/category/reset/<int:cat_id>', methods=['POST'])
@login_required
def admin_reset_category(cat_id):
    db = get_db(); cat = db.execute("SELECT * FROM categories WHERE id=?", (cat_id,)).fetchone()
    if not cat: flash('Not found'); return redirect(url_for('admin_funds'))
    fids = [f['id'] for f in db.execute("SELECT id FROM funds WHERE category_id=?", (cat_id,)).fetchall()]
    if fids:
        ph = ','.join('?' * len(fids))
        for tbl in ['navs', 'precalc_returns', 'fund_details', 'fund_details_history']:
            db.execute(f"DELETE FROM {tbl} WHERE fund_id IN ({ph})", fids)
        db.execute("DELETE FROM funds WHERE category_id=?", (cat_id,))
    db.commit(); flash(f'Category "{cat["name"]}" fully reset')
    return redirect(url_for('admin_funds'))

@app.route('/admin/category/delete_navs/<int:cat_id>', methods=['POST'])
@login_required
def admin_delete_navs(cat_id):
    db = get_db(); cat = db.execute("SELECT * FROM categories WHERE id=?", (cat_id,)).fetchone()
    if not cat: flash('Not found'); return redirect(url_for('admin_funds'))
    fids = [f['id'] for f in db.execute("SELECT id FROM funds WHERE category_id=?", (cat_id,)).fetchall()]
    if fids:
        ph = ','.join('?' * len(fids))
        db.execute(f"DELETE FROM navs WHERE fund_id IN ({ph})", fids)
        db.execute(f"DELETE FROM precalc_returns WHERE fund_id IN ({ph})", fids)
        db.execute("UPDATE funds SET qualifies_rolling=0, inception_date=NULL WHERE category_id=?", (cat_id,))
    db.commit(); flash(f'NAV history deleted for "{cat["name"]}"')
    return redirect(url_for('admin_funds'))

@app.route('/admin/category/delete_details/<int:cat_id>', methods=['POST'])
@login_required
def admin_delete_details(cat_id):
    db = get_db(); cat = db.execute("SELECT * FROM categories WHERE id=?", (cat_id,)).fetchone()
    if not cat: flash('Not found'); return redirect(url_for('admin_funds'))
    fids = [f['id'] for f in db.execute("SELECT id FROM funds WHERE category_id=?", (cat_id,)).fetchall()]
    if fids:
        ph = ','.join('?' * len(fids))
        db.execute(f"DELETE FROM fund_details WHERE fund_id IN ({ph})", fids)
        db.execute(f"DELETE FROM fund_details_history WHERE fund_id IN ({ph})", fids)
    db.commit(); flash(f'Fund details deleted for "{cat["name"]}"')
    return redirect(url_for('admin_funds'))

# ── ADMIN ──────────────────────────────────────────────────────────────────────
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if check_admin_pwd(request.form.get('password', '')):
            session.permanent = True
            session['admin_logged_in'] = True; return redirect(url_for('admin_funds'))
        flash('Incorrect password')
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout(): session.pop('admin_logged_in', None); return redirect(url_for('index'))

@app.route('/admin/funds')
@login_required
def admin_funds():
    db = get_db()
    return render_template('admin_funds.html',
        funds=db.execute("SELECT f.*,c.name as category_name FROM funds f JOIN categories c ON c.id=f.category_id ORDER BY c.display_order,f.type DESC,f.name").fetchall(),
        categories=db.execute("SELECT * FROM categories WHERE is_active=1 ORDER BY display_order").fetchall())

@app.route('/admin/funds/add', methods=['POST'])
@login_required
def admin_add_fund():
    db = get_db()
    name, cid = request.form.get('name', '').strip(), request.form.get('category_id', type=int)
    ft, inc = request.form.get('type', 'fund'), request.form.get('inception_date', '')
    if not name or not cid: flash('Name and category required'); return redirect(url_for('admin_funds'))
    try:
        short = name.replace(' Fund-Reg(G)', '').replace(' Fund(G)', '')
        db.execute("INSERT INTO funds(category_id,name,short_name,type,inception_date) VALUES(?,?,?,?,?)",
                   (cid, name, short, ft, inc or None))
        db.execute("INSERT OR IGNORE INTO fund_details(fund_id) SELECT id FROM funds WHERE name=? AND category_id=?",
                   (name, cid))
        db.commit(); flash(f'Fund "{name}" added')
    except Exception as e: flash(f'Error: {e}')
    return redirect(url_for('admin_funds'))

@app.route('/admin/funds/toggle/<int:fund_id>', methods=['POST'])
@login_required
def admin_toggle_fund(fund_id):
    db = get_db()
    cur = db.execute("SELECT is_active FROM funds WHERE id=?", (fund_id,)).fetchone()
    if cur:
        db.execute("UPDATE funds SET is_active=? WHERE id=?", (0 if cur['is_active'] else 1, fund_id))
        db.commit()
    return redirect(url_for('admin_funds'))

@app.route('/admin/funds/delete/<int:fund_id>', methods=['POST'])
@login_required
def admin_delete_fund(fund_id):
    db = get_db()
    fund = db.execute("SELECT name FROM funds WHERE id=?", (fund_id,)).fetchone()
    if fund:
        db.execute("UPDATE funds SET is_active=0 WHERE id=?", (fund_id,))
        db.execute("DELETE FROM precalc_returns WHERE fund_id=?", (fund_id,))
        db.execute("DELETE FROM fund_details WHERE fund_id=?", (fund_id,))
        db.commit(); flash(f'Fund "{fund["name"]}" deleted (NAVs archived)')
    return redirect(url_for('admin_funds'))

@app.route('/admin/nav_history')
@login_required
def nav_history():
    return render_template('nav_history.html',
        categories=get_db().execute("SELECT * FROM categories WHERE is_active=1 ORDER BY display_order").fetchall())

@app.route('/admin/fund_details')
@login_required
def admin_fund_details_page():
    return render_template('admin_fund_details.html',
        categories=get_db().execute("SELECT * FROM categories WHERE is_active=1 ORDER BY display_order").fetchall())

@app.route('/admin/fund_details_hybrid')
@login_required
def admin_fund_details_hybrid_page():
    return render_template('admin_fund_details_hybrid.html',
        categories=get_db().execute("SELECT * FROM categories WHERE is_active=1 ORDER BY display_order").fetchall())

if __name__ == '__main__':
    app.run(debug=True, port=5000)
